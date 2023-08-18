# Parse survey excel and create pdf for each response

import os

from functools import partial
from pathlib import Path

from openpyxl import load_workbook, Workbook
from openpyxl.utils.escape import escape, unescape

from reportlab.platypus import BaseDocTemplate, Frame, Image, PageTemplate, Paragraph
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import mm
from reportlab.lib.enums import TA_RIGHT
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont

# Arial fonts
pdfmetrics.registerFont(TTFont('Arial', 'Arial.ttf'))
pdfmetrics.registerFont(TTFont('Arial-B', 'Arial Bold.ttf'))
pdfmetrics.registerFont(TTFont('Arial-I', 'Arial Italic.ttf'))
pdfmetrics.registerFont(TTFont('Arial-N', 'Arial Narrow.ttf'))

# Add required styles for the texts
styles = getSampleStyleSheet()
# Report title
styles.add(
    ParagraphStyle(
        name="ntitle",
        fontName="Arial-B",
        fontSize=14,
        textColor="#000000",
        leading=16
    )
)
# User name
styles.add(
    ParagraphStyle(
        name="name",
        fontName="Arial",
        fontSize=11,
        textColor="#000000"
    )
)
# User email
styles.add(
    ParagraphStyle(
        name="email",
        fontName="Arial-I",
        fontSize=11,
        textColor="#2D2D2D"
    )
)
# Disclaimer text for multi platform
styles.add(
    ParagraphStyle(
        name="multi-plt",
        fontName="Arial-I",
        fontSize=9,
        textColor="#2D2D2D",
        alignment=TA_RIGHT,
        leading=10
    )
)
# Platform in header (technology)
styles.add(
    ParagraphStyle(
        name="technology-plt",
        fontName="Arial-B",
        fontSize=11,
        textColor="#4C979F",
        alignment=TA_RIGHT,
        leading=12
    )
)
# Non exisitng platform in header (technology)
styles.add(
    ParagraphStyle(
        name="technology-non-plt",
        fontName="Arial-B",
        fontSize=10,
        textColor="#4C979F",
        alignment=TA_RIGHT,
        leading=11
    )
)
# Platform in header (facility)
styles.add(
    ParagraphStyle(
        name="facility-plt",
        fontName="Arial-B",
        fontSize=11,
        textColor="#A7C947",
        alignment=TA_RIGHT,
        leading=12
    )
)
# Non exisitng platform in header (facility)
styles.add(
    ParagraphStyle(
        name="facility-non-plt",
        fontName="Arial-B",
        fontSize=10,
        textColor="#A7C947",
        alignment=TA_RIGHT,
        leading=11
    )
)
# Question/section (technology)
styles.add(
    ParagraphStyle(
        name="technology",
        fontName="Arial-B",
        fontSize=11,
        textColor="#4C979F",
        spaceAfter=5,
        leading=13
    )
)
# Question/section (facility)
styles.add(
    ParagraphStyle(
        name="facility",
        fontName="Arial-B",
        fontSize=11,
        textColor="#A7C947",
        spaceAfter=5,
        leading=13
    )
)
# Normal texts
styles.add(
    ParagraphStyle(
        name="normal",
        fontName="Arial",
        fontSize=11,
        textColor="#2D2D2D",
        spaceAfter=10,
        leading=14
    )
)
# Footer texts
styles.add(
    ParagraphStyle(
        name="footer",
        fontName="Arial",
        fontSize=10,
        textColor="#A6ACAF"
    )
)

class report_gen(object):
    # A class object that defines the layout of pdf
    def __init__(self, filename):
        self.filename = filename
        self.doc = BaseDocTemplate(
                    self.filename,
                    pagesize=A4,
                    rightMargin=14 * mm,
                    leftMargin=14 * mm,
                    topMargin=12 * mm,
                    bottomMargin=12 * mm,
                    showBoundary=0,
                   )
        self.__header_content = []
        self.__footer_content = []
        self.__content = []
    
    # Add the text to header section
    def add_to_header(self, text, style):
        self.__header_content.append(Paragraph(text, style))
    # Add the text to footer section
    def add_to_footer(self, text, style=None):
        if style:
            self.__footer_content.append(Paragraph(text, style))
        else:
            self.__footer_content.append(text)
    # Add the text to main section
    def add_to_content(self, text, style):
        self.__content.append(Paragraph(text, style))
    # Function will make the page layout and generate pdf
    def make_pdf(self):
        # make page layouts
        self.__make_page_layout()
        # create the directories
        Path(os.path.split(self.filename)[0]).mkdir(parents=True, exist_ok=True)
        # make the pdf
        self.doc.build(self.__content)
    
    # This funtion creates the page layout by creating main body frame
    # and exisiting header function (which defines header layout)
    def __make_page_layout(self):
        header_height = self.__get_header_height()
        col1 = Frame(
            id="col1",
            x1=self.doc.leftMargin,
            y1=self.doc.bottomMargin + 6*mm,
            width=self.doc.width/2 - 3*mm,
            height=self.doc.height - header_height,
            leftPadding=0 * mm,
            topPadding=2 * mm,
            rightPadding=0 * mm,
            bottomPadding=2 * mm,
            #showBoundary=0.5
        )
        col2 = Frame(
            id="col2",
            x1=self.doc.leftMargin + self.doc.width/2 + 3*mm,
            y1=self.doc.bottomMargin + 6*mm,
            width=self.doc.width/2 - 3*mm,
            height=self.doc.height - header_height,
            leftPadding=0 * mm,
            topPadding=2 * mm,
            rightPadding=0 * mm,
            bottomPadding=2 * mm,
            #showBoundary=0.5
        )
        # Add the above created frames to a template
        template = PageTemplate(id="all_frames", frames=[col1, col2], onPage=self.__call_header_and_footer)
        # Add the above temple to the base doc
        self.doc.addPageTemplates([template])
        
    # Function that controls layout of header
    def __header(self, canvas, doc, content):
        canvas.saveState()
        page_num = canvas.getPageNumber()
        # Write title of the document, add a tag for multi page
        if page_num == 2:
            content[0].frags[0].text = content[0].frags[0].text + " (cont.)"
        w0, h0 = content[0].wrap(doc.width * 0.7, doc.topMargin)
        content[0].drawOn(canvas, doc.leftMargin, doc.height + doc.topMargin - h0)
        # Write the name of the user
        w1, h1 = content[1].wrap(doc.width * 0.7, doc.topMargin + h0)
        h1 = h1 + h0 + 2*mm
        content[1].drawOn(canvas, doc.leftMargin, doc.height + doc.topMargin - h1)
        # Write the email address
        w2, h2 = content[2].wrap(doc.width * 0.7, doc.topMargin + h1)
        h2 = h2 + h1 + 1*mm
        content[2].drawOn(canvas, doc.leftMargin, doc.height + doc.topMargin - h2)
        # Write the platform title
        w3, h3 = content[3].wrap(doc.width, doc.topMargin)
        content[3].drawOn(canvas, doc.leftMargin, doc.height + doc.topMargin - h3)
        # Write notification if proposal multiple platform
        if len(content) == 5:
            w4, h4 = content[4].wrap(doc.width, doc.topMargin + h0)
            h4 = h4 + h0 + 3*mm
            content[4].drawOn(canvas, doc.leftMargin, doc.height + doc.topMargin - h4)
        # Draw a horizantal divider line
        hl = h2 + 2*mm
        p = canvas.beginPath()
        p.moveTo(doc.leftMargin, doc.height + doc.topMargin - hl)
        p.lineTo(doc.leftMargin + doc.width, doc.height + doc.topMargin - hl)
        p.close()
        canvas.setLineWidth(0.5)
        canvas.setStrokeColor("#999999")
        canvas.drawPath(p, stroke=1)
        canvas.restoreState()
    
    # Function that controls layout of footer
    def __footer(self, canvas, doc, content):
        canvas.saveState()
        # Put reg num
        w0, h0 = content[0].wrap(doc.width, doc.bottomMargin)
        content[0].drawOn(canvas, doc.leftMargin, doc.bottomMargin)
        # Draw logo
        content[1].drawOn(canvas, doc.leftMargin + doc.width - 22*mm, doc.bottomMargin)
        p = canvas.beginPath()
        p.moveTo(doc.leftMargin, doc.bottomMargin + 6*mm)
        p.lineTo(doc.leftMargin + doc.width, doc.bottomMargin + 6*mm)
        p.close()
        canvas.setLineWidth(0.5)
        canvas.setStrokeColor("#999999")
        canvas.drawPath(p, stroke=1)
        canvas.restoreState()
    
    # Wrapper function to call both header and footer
    def __call_header_and_footer(self, canvas, doc):
        self.__header(canvas, doc, self.__header_content)
        self.__footer(canvas, doc, self.__footer_content)
    
    # Function to calculate header height
    def __get_header_height(self):
        w0, h0 = self.__header_content[0].wrap(self.doc.width * 0.7, self.doc.topMargin)
        w1, h1 = self.__header_content[1].wrap(self.doc.width * 0.7, self.doc.topMargin + h0)
        h1 = h1 + h0 + 2*mm
        w2, h2 = self.__header_content[2].wrap(self.doc.width * 0.7, self.doc.topMargin + h1)
        return h2 + h1 + 8*mm

# Helper method for aesthetic
def platform_outside_scilifelab(platform):
    return platform in ["None of the existing platforms", "None of the current platforms", "I do not know", "No platform suggested"]

def get_platform_header_text(platform, plt_text):
    # If its not an exisitng platform, put in special criteria
    if platform_outside_scilifelab(platform):
        if plt_text == "technology":
            text = "Proposals on Technologies with<br/>no specific platform suggested"
        else:
            text = "Proposals on Infrastructure Units with<br/>no specific platform suggested"
    elif platform == "Clinical Proteomics and Immunology":
        text = "Clinical Proteomics<br/>and Immunology"
    elif platform == "Cellular and Molecular Imaging":
        text = "Cellular and Molecular<br/>Imaging"
    elif platform == "Integrated Structural Biology":
        text = "Integrated Structural<br/>Biology"
    elif platform == "Chemical Biology and Genome Engineering":
        text = "Chemical Biology and<br/>Genome Engineering"
    elif platform == "Drug Discovery and Development":
        text = "Drug Discovery and<br/>Development"
    else:
        text = platform
    return text

# Info dict regarding survey and styling specs whihc are later used downstream
suggestions_info = {
    "A": {
        "alt_text": "technology",
        "plt_text": "technology",
        "footer_text": "Proposal on new SciLifeLab Technology",
        "style": "technology",
        "style_plt": "technology-plt",
        "style_non_plt": "technology-non-plt",
        "title_index": 10,
        "platform_index": 12,
        "reg_num": 0
    },
    "B": {
        "alt_text": "facility",
        "plt_text": "Scilifelab unit",
        "footer_text": "Proposal on new SciLifeLab Unit",
        "style": "facility",
        "style_plt": "facility-plt",
        "style_non_plt": "facility-non-plt",
        "title_index": 18,
        "platform_index": 26,
        "reg_num": 0
    }
}

platforms_order = ["Bioinformatics", "Genomics", "Clinical Genomics", "Clinical Proteomics and Immunology",
                   "Metabolomics", "Spatial Biology", "Cellular and Molecular Imaging", "Integrated Structural Biology",
                   "Chemical Biology and Genome Engineering", "Drug Discovery and Development", "No platform suggested"]

# Read the survey file
wb = load_workbook('Survey.xlsx')
ws = wb.active
rpn = 0
ntotal = 0

# excel file for metadata
owb = Workbook()
ows = owb.active
for c, h in enumerate(["Report Num.", "Reg Num.", "Title", "Platform", "Category"], 1):
    cell = ows.cell(row=1, column=c)
    cell.value = h

# Read and sort the data to process in right order
process_order = {}
reg_num = {"A": 1, "B": 1}
for n, srow in enumerate(ws.iter_rows(min_row=3), 3):
    row = [unescape(str(cell.value) or "") for cell in srow]
    sid = row[9][0].upper()
    platforms = [p.strip() for p in row[suggestions_info[sid]["platform_index"]].split(", ")]
    platforms_uniq = list(set(["No platform suggested" if platform_outside_scilifelab(p) else p for p in platforms]))
    for p in platforms_uniq:
        if p not in process_order:
            process_order[p] = {"A": [], "B": []}
        s_title = row[suggestions_info[sid]["title_index"]].strip()
        process_order[p][sid].append({
            "title": s_title[0].upper() + s_title[1:],
            "reg_no": sid + str(reg_num[sid]),
            "row_index": n,
            "multi_platform": len(platforms_uniq) > 1
        })
        ntotal += 1
    reg_num[sid] += 1

for plt_i, p in enumerate(platforms_order, 1):
    if p not in process_order:
        continue
    for i in ["A", "B"]:
        for s in sorted(process_order[p][i], key=lambda d: d['title'].lower()):
            row = [unescape(str(cell.value) or "") for cell in ws[s["row_index"]]]
            rpn += 1
            rpid = str(rpn).zfill(len(str(ntotal)))
            snm = suggestions_info[i]["style"]
            snm_plt = suggestions_info[i]["style_plt"]
            snm_non_plt = suggestions_info[i]["style_non_plt"]
            platforms = [p.strip() for p in row[suggestions_info[i]["platform_index"]].split(", ")]
            # Filename and path
            pdf_name = "{}_{}_{}.pdf".format(rpid, s["title"].replace(" ", "_"), s["reg_no"])
            fname = os.path.join("Pdfs", "{}_{}".format(str(plt_i), p), pdf_name)
            # Instantiate report gen object
            rp = report_gen(fname)
            # Affiliation text
            if row[4] == "University":
                aff_text = row[6]
            elif row[7] == "None":
                aff_text = row[4]
            else:
                aff_text = "{}, {}".format(row[4], row[7])
            # Add content to header section
            rp.add_to_header("{}: {}".format(rpid, s["title"]), styles["ntitle"])
            rp.add_to_header("{} {}, {}, {}".format(row[0], row[1], row[2], aff_text), styles["name"])
            rp.add_to_header(row[3], styles["email"])
            rp.add_to_header(get_platform_header_text(p, suggestions_info[i]["plt_text"]), styles[snm_plt])
            # Add disclaimer if proposal belongs to two platform
            if s["multi_platform"]:
                rp.add_to_header(
                        "**Please note that this proposal is<br/>also found under other platforms",
                        styles["multi-plt"]
                    )
            # Add content to Footer
            rp.add_to_footer("{} - Report No: {}, Reg No: {}".format(suggestions_info[i]["footer_text"], rpid, s["reg_no"]), styles["footer"])
            rp.add_to_footer(Image("SciLifeLab_logo.png", width=22*mm, height=5*mm))
            # Add content to main body
            # Representing text
            if row[5] == "Other":
                rep_text = row[8]
            elif row[8] == "None":
                rep_text = row[5]
            else:
                rep_text = "{} ({})".format(row[5], row[8])
            rp.add_to_content("Representing:", styles[snm])
            rp.add_to_content(rep_text, styles["normal"])
            # Platforms text
            rp.add_to_content("The {} would fit in the SciLifeLab Platform(s):".format(suggestions_info[i]["alt_text"]), styles[snm])
            rp.add_to_content("<br/>".join(platforms), styles["normal"])
            # Info relavant for technology/service proposal
            if i == "A":
                # Contribution to scilifelab or ddls
                rp.add_to_content("The suggested technology would contribute to following capabilities:", styles[snm])
                rp.add_to_content(row[13].replace(", ", "<br/>"), styles["normal"])
                # Currently available
                if row[14] == "No" or row[15] == "None":
                    avail_text = row[14]
                else:
                    avail_text = "{}, {}".format(row[14], row[15])
                rp.add_to_content("Is the technology currently available as local infrastructure service in Sweden?", styles[snm])
                rp.add_to_content(avail_text, styles["normal"])
                # Brief description
                rp.add_to_content("Brief description of the technology:", styles[snm])
                rp.add_to_content(row[11].replace("\n", "<br/>"), styles["normal"])
                # Estimated funding
                rp.add_to_content("Estimated annual total funding (MSEK) needed from SciLifeLab:", styles[snm])
                rp.add_to_content(row[16].replace("\n", "<br/>"), styles["normal"])
                # Additional comment
                rp.add_to_content("Additional comment:", styles[snm])
                rp.add_to_content(row[17].replace("\n", "<br/>"), styles["normal"])
            # Info relavant for facility proposal
            else:
                # Facility location
                rp.add_to_content("Facility location:", styles[snm])
                rp.add_to_content(row[19], styles["normal"])
                # Contact person name
                rp.add_to_content("Contact person for the facility:", styles[snm])
                rp.add_to_content(row[20], styles["normal"])
                # Contact person email
                rp.add_to_content("Contact person email address:", styles[snm])
                rp.add_to_content(row[21], styles["normal"])
                # Uniq users
                rp.add_to_content("Current number of unique users annually:", styles[snm])
                rp.add_to_content(row[25], styles["normal"])
                # Contribution to scilifelab or ddls
                rp.add_to_content("The suggested facility would contribute to following capabilities:", styles[snm])
                rp.add_to_content(row[27].replace(", ", "<br/>"), styles["normal"])
                # Uniq users estimate
                rp.add_to_content("Estimated unique annual users if the unit become a part of SciLifeLab infrastructure:", styles[snm])
                rp.add_to_content(row[28], styles["normal"])
                # Brief description
                rp.add_to_content("Brief description of the facility:", styles[snm])
                rp.add_to_content(row[22].replace("\n", "<br/>"), styles["normal"])
                # Services providing today
                if row[23] == "I do not know" or row[24] == "None":
                    provide_text = row[23]
                else:
                    provide_text = "{}, {}".format(row[23], row[24])
                rp.add_to_content("How is the facility providing infrastructure services today?", styles[snm])
                rp.add_to_content(provide_text, styles["normal"])
                # Estimated funding
                rp.add_to_content("Estimated annual funding (MSEK) needed from SciLifeLab, co-funding and user fee plans:", styles[snm])
                rp.add_to_content(row[29].replace("\n", "<br/>"), styles["normal"])
                # Additional comment
                rp.add_to_content("Additional comment:", styles[snm])
                rp.add_to_content(row[30].replace("\n", "<br/>"), styles["normal"])
            rp.make_pdf()
            # following is to generate meta data
            for c, v in enumerate([rpid, s["reg_no"], s["title"], p, "Technology" if i == "A" else "Unit"], 1):
                cell = ows.cell(row=rpn+1, column=c)
                cell.value = v
owb.save("Survey_meta.xlsx")
